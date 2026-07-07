#!/usr/bin/env python3
"""Import 'Overview visits.xlsx' into the Atlas Supabase database.

Reads the Visits sheet (ongoing deals only) and the Dropdown list sheet
(canonical projects, buildings, broker firms), then REPLACES the assets,
brokers, stages and leads tables. Managers are left untouched.

Usage:
  python3 scripts/import_visits.py <workbook.xlsx>           # dry run: print summary + write import.sql
  python3 scripts/import_visits.py <workbook.xlsx> --apply   # also apply to Supabase via Management API

--apply reads the access token from ~/.supabase/access-token.
"""
import json
import pathlib
import re
import subprocess
import sys
from datetime import date, datetime

PROJECT_REF = 'pcrhodfcsgkexckfrnak'

import openpyxl


def norm(s):
    return re.sub(r'\W+', '', str(s)).lower()


def slug(s):
    return re.sub(r'[^a-z0-9]+', '-', str(s).strip().lower()).strip('-') or 'x'


def esc(s):
    return str(s).replace("'", "''")


def sqlval(v):
    if v is None:
        return 'null'
    if isinstance(v, bool):
        return 'true' if v else 'false'
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, (dict, list)):
        return "'" + esc(json.dumps(v, ensure_ascii=False)) + "'::jsonb"
    return "'" + esc(v) + "'"


def as_date(v):
    if v is None or v == '':
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(s[:10], fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    return None


def as_text(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    return s or None


def as_sqm(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return float(v)
    m = re.search(r'\d[\d.,]*', str(v))
    if not m:
        return 0
    try:
        return float(m.group().replace('.', '').replace(',', '.'))
    except ValueError:
        return 0


def initials(name):
    parts = str(name).strip().split()
    return ''.join(p[0] for p in parts[:2]).upper() or '??'


def short(name, n=12):
    name = str(name).strip()
    return name if len(name) <= n else name[:n] + '…'


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    path = sys.argv[1]
    apply_it = '--apply' in sys.argv

    wb = openpyxl.load_workbook(path, data_only=True)

    # ---- canonical vocab from the Dropdown sheet
    dd = wb['Dropdown list']
    grid = [[c for c in row] for row in dd.iter_rows(values_only=True)]

    # column A: project list — stop at the first blank (other lists are stacked below)
    projects = []
    for row in grid[1:]:
        v = as_text(row[0])
        if not v:
            break
        projects.append(v)

    # building columns: header row has project names from col C onward until ''
    buildings_by_project = {}
    header = grid[0]
    for ci in range(2, len(header)):
        pname = as_text(header[ci])
        if not pname or norm(pname) == 'broker':
            break
        subs = []
        for row in grid[1:]:
            v = as_text(row[ci]) if ci < len(row) else None
            if v:
                subs.append(v)
        buildings_by_project[norm(pname)] = subs

    # broker firm list: column headed 'Broker'
    broker_col = next(i for i, v in enumerate(header) if v and norm(v) == 'broker')
    firm_names = []
    for row in grid[1:]:
        v = as_text(row[broker_col]) if broker_col < len(row) else None
        if v:
            firm_names.append(v)

    # ---- assets
    assets = {}  # key: norm name
    def ensure_asset(name):
        k = norm(name)
        if k in assets:
            return assets[k]
        display = name.strip()
        a = {
            'id': slug(display), 'name': display, 'short': short(display),
            'type': 'Retail' if 'retail' in k else 'Office',
            'loc': '—', 'manager': 'LV', 'tenant_rep': None, 'subs': [],
        }
        for bname in buildings_by_project.get(k, []):
            a['subs'].append({
                'id': slug(bname), 'name': bname.strip(), 'short': short(bname, 10),
                'sqm': 0, 'occ': 0, 'units': 0, 'vacant': 0, 'rent': 0,
            })
        assets[k] = a
        return a

    for p in projects:
        ensure_asset(p)

    # ---- brokers
    brokers = {}  # key: norm firm
    def ensure_broker(name):
        k = norm(name)
        if k in brokers:
            return brokers[k]
        b = {'id': slug(name), 'name': str(name).strip(), 'contacts': []}
        brokers[k] = b
        return b

    for f in firm_names:
        ensure_broker(f)

    # ---- visits sheet
    ws = wb['Visits']
    rows = list(ws.iter_rows(values_only=True))
    hdr_i = next(i for i, r in enumerate(rows) if r[1] and norm(r[1]) == 'status')
    hdr = {norm(v): i for i, v in enumerate(rows[hdr_i]) if v}
    col = lambda name: hdr[norm(name)]

    leads = []
    skipped = 0
    for r in rows[hdr_i + 1:]:
        status = as_text(r[col('Status')])
        company = as_text(r[col('Company name')])
        if not company or not status or norm(status) != 'ongoing':
            continue
        project = as_text(r[col('Project')])
        if not project:
            skipped += 1
            continue
        a = ensure_asset(project)

        # building → sub (add unseen buildings so nothing is lost)
        sub_id = None
        bname = as_text(r[col('Building')])
        if bname:
            bk = norm(bname)
            match = next((s for s in a['subs'] if norm(s['name']) == bk or bk in norm(s['name']) or norm(s['name']) in bk), None)
            if not match:
                match = {'id': slug(bname), 'name': bname, 'short': short(bname, 10),
                         'sqm': 0, 'occ': 0, 'units': 0, 'vacant': 0, 'rent': 0}
                a['subs'].append(match)
            sub_id = match['id']

        # broker firm + contact
        broker_id = contact_id = None
        firm = as_text(r[col('Broker')])
        if firm:
            b = ensure_broker(firm)
            broker_id = b['id']
            person = as_text(r[col('Broker contact')])
            if person:
                c = next((c for c in b['contacts'] if norm(c['name']) == norm(person)), None)
                if not c:
                    c = {'id': slug(person), 'name': person, 'init': initials(person)}
                    b['contacts'].append(c)
                contact_id = c['id']

        intro = as_date(r[col('Introduction')])
        visits = [d for d in (as_date(r[col(f'Visit {i}')]) for i in (1, 2, 3, 4)) if d]
        last_proposal = as_date(r[col('Last Proposal')])
        agreed = as_date(r[col('Proposal agreed')])

        if agreed:
            stage = 'agreed'
        elif last_proposal:
            stage = 'proposal'
        elif visits:
            stage = 'visited'
        elif intro:
            stage = 'intro'
        else:
            stage = 'new'

        kind_raw = norm(as_text(r[col('Current tenant or new tenant')]) or 'new tenant')
        comments = as_text(r[col('Comments')])
        leads.append({
            'id': f'v{len(leads) + 1}',
            'company': company, 'contact': '',
            'type': a['type'], 'sqm': as_sqm(r[col('Sqm')]),
            'asset_id': a['id'], 'sub_id': sub_id, 'stage': stage,
            'broker': broker_id, 'broker_contact': contact_id,
            'tenant_kind': 'current' if 'current' in kind_raw else 'new',
            'deal_type': as_text(r[col('Renewal/ Extension/ Reduction')]),
            'activity': as_text(r[col('Activity')]) or as_text(r[col('Activity (Nacebel)')]),
            'timing': as_text(r[col('Timing')]),
            'intro_date': intro, 'visits': visits,
            'last_proposal': last_proposal, 'proposal_agreed': agreed,
            'comments': comments, 'next_step': comments or '',
        })

    stages = [
        ('new', 'New', '#948A7B'), ('intro', 'Introduced', '#B08327'),
        ('visited', 'Visited', '#C05F2E'), ('proposal', 'Proposal', '#9D4A26'),
        ('agreed', 'Agreed', '#74803B'), ('signed', 'Signed', '#4C8355'),
        ('out', 'Out', '#8A8578'),
    ]

    # ---- SQL
    out = ['begin;', 'delete from public.leads;', 'delete from public.stages;',
           'delete from public.assets;', 'delete from public.brokers;']
    for i, (sid, label, dot) in enumerate(stages):
        out.append(f"insert into public.stages (id,label,dot,position) values ({sqlval(sid)},{sqlval(label)},{sqlval(dot)},{i});")
    for i, a in enumerate(assets.values()):
        out.append(
            'insert into public.assets (id,name,short,type,loc,manager,tenant_rep,subs,position) values ('
            + ','.join(sqlval(v) for v in (a['id'], a['name'], a['short'], a['type'], a['loc'], a['manager'], a['tenant_rep'], a['subs'], i)) + ');')
    for i, b in enumerate(brokers.values()):
        out.append(
            'insert into public.brokers (id,name,contacts,position) values ('
            + ','.join(sqlval(v) for v in (b['id'], b['name'], b['contacts'], i)) + ');')
    lead_cols = ['id', 'company', 'contact', 'type', 'sqm', 'asset_id', 'sub_id', 'stage', 'broker',
                 'broker_contact', 'tenant_kind', 'deal_type', 'activity', 'timing', 'intro_date',
                 'visits', 'last_proposal', 'proposal_agreed', 'comments', 'next_step']
    for i, l in enumerate(leads):
        vals = [l[c] for c in lead_cols] + [i]
        out.append(f"insert into public.leads ({','.join(lead_cols)},position) values ("
                   + ','.join(sqlval(v) for v in vals) + ');')
    out.append('commit;')
    sql = '\n'.join(out)
    pathlib.Path('scripts/import.sql').write_text(sql)

    n_contacts = sum(len(b['contacts']) for b in brokers.values())
    n_subs = sum(len(a['subs']) for a in assets.values())
    from collections import Counter
    print(f"assets: {len(assets)} ({n_subs} buildings)")
    print(f"brokers: {len(brokers)} ({n_contacts} contacts)")
    print(f"leads (ongoing): {len(leads)}  skipped (no project): {skipped}")
    print('stage distribution:', dict(Counter(l['stage'] for l in leads)))
    print('tenant kind:', dict(Counter(l['tenant_kind'] for l in leads)))
    print('sql statements:', len(out), '→ scripts/import.sql')

    if apply_it:
        token = pathlib.Path.home().joinpath('.supabase/access-token').read_text().strip()
        body = pathlib.Path('scripts/import.body.json')
        body.write_text(json.dumps({'query': sql}))
        r = subprocess.run(
            ['curl', '-s', '-w', '\nHTTP %{http_code}', '-X', 'POST',
             f'https://api.supabase.com/v1/projects/{PROJECT_REF}/database/query',
             '-H', f'Authorization: Bearer {token}', '-H', 'Content-Type: application/json',
             '--data', f'@{body}'],
            capture_output=True, text=True)
        body.unlink()
        print('apply:', r.stdout[-200:])


if __name__ == '__main__':
    main()
